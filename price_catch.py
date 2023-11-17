import xlrd
import openpyxl

# 打開原始Excel檔案（使用xlrd來讀取舊的.xls格式）
input_file = "蔬菜產品行情比較.xls"
workbook = xlrd.open_workbook(input_file)
sheet = workbook.sheet_by_index(0)

# 要提取的欄位範圍
columns_to_extract = [0, 2, 3]
start_row = 6  # 注意：xlrd的索引從0開始，所以這裡使用7表示第8行
end_row = sheet.nrows

# 創建一個新的Excel檔案
output_file = "output.xlsx"
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# 提取資料並寫入新的Excel檔案
for row in range(start_row, end_row):
    output_row = []
    for column_index in columns_to_extract:
        cell_value = sheet.cell_value(row, column_index)
        output_row.append(cell_value)

    output_sheet.append(output_row)

# 保存新的Excel文件
output_workbook.save(output_file)

##############

# 打开Excel文件
input_file = "output.xlsx"
output_file = "output.xlsx"  # 输出文件与输入文件相同

# 打开工作簿
workbook = openpyxl.load_workbook(input_file)
sheet = workbook.active

# 从B列和C列获取数据并四舍五入
for row in sheet.iter_rows(min_row=1, min_col=2, max_col=3):
    for cell in row:
        if cell.value is not None:
            cell.value = round(cell.value)  # 四舍五入为整数

# 保存修改后的工作簿
workbook.save(output_file)
print("資料處理完成")

##################

import openpyxl

# 打开Excel文件
input_file = "output.xlsx"
output_file = "output.xlsx"  # 输出文件与输入文件相同

# 打开工作簿
workbook = openpyxl.load_workbook(input_file)
sheet = workbook.active

# 插入一列在A列之后
sheet.insert_cols(2)  # 在A列之后插入一列

# 复制A列的数据到新插入的列
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        if cell.value is not None:
            new_cell = sheet.cell(row=cell.row, column=2, value=cell.value)

# 保存修改后的工作簿
workbook.save(output_file)

####################

import openpyxl
import re

# 打开Excel文件
input_file = "output.xlsx"
output_file = "output.xlsx"  # 输出文件与输入文件相同

# 打开工作簿
workbook = openpyxl.load_workbook(input_file)
sheet = workbook.active

# 处理A列：保留英文和数字，删除中文
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        if cell.value is not None:
            # 使用正则表达式匹配中文字符并删除
            cell.value = re.sub(r'[\u4e00-\u9fff]', '', cell.value)

# 处理B列：保留中文，删除英文和数字
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=2):
    for cell in row:
        if cell.value is not None:
            # 使用正则表达式匹配英文和数字并删除
            cell.value = re.sub(r'[a-zA-Z0-9]', '', cell.value)

sheet.delete_cols(4)
# 保存修改后的工作簿
workbook.save(output_file)

###

import csv
import openpyxl

# 打开output.xlsx
input_file = "output.xlsx"
workbook = openpyxl.load_workbook(input_file)
sheet = workbook.active

# 创建一个CSV文件并打开以写入
output_csv = "price.csv"
with open(output_csv, 'w', newline='', encoding='utf-8') as csv_file:
    csv_writer = csv.writer(csv_file)

    # 遍历工作表中的行
    for row in sheet.iter_rows(values_only=True):
        # 使用列表推导式去除每个字段中的空格
        cleaned_row = [str(cell).strip() for cell in row]
        csv_writer.writerow(cleaned_row)

print(f"已轉為 {output_csv}")

import os

file_to_delete1 = "蔬菜產品行情比較.xls"
file_to_delete2 = "output.xlsx"

os.remove(file_to_delete1)
os.remove(file_to_delete2)
